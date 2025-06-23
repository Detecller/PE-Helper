import discord
from discord.ext import commands
from discord import app_commands, Object
from utils.permissions import has_allowed_role_and_channel
from utils.variables import SGT
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
from utils.setup_logger import log_slash_command
import logging


GUILD_ID = int(os.getenv("GUILD_ID"))

# Get logger
logger = logging.getLogger("pe_helper")


class EXCOExclusive(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot

    exco_exclusive_group = app_commands.Group(name="exco-exclusive", description="EXCO-related commands")

    @exco_exclusive_group.command(name="members-details", description="Exports member & alumni details to Excel.")
    @has_allowed_role_and_channel(allowed_roles=['Current EXCO', 'Admin'], allowed_channels=['👑┃exco-exclusive', '⚙️┃admin-related'])
    async def members_details(self, interaction: discord.Interaction):

        await interaction.response.defer()

        log_slash_command(logger, interaction)
        
        guild = interaction.guild
        rows = []
        for m in guild.members:
            if m.bot: continue
            roles = {r.name for r in m.roles}

            # Assign role conditionally
            if roles & {"Member","Alumni","Current EXCO"}:
                if "Current EXCO" in roles:
                    status = "Current EXCO"
                elif "Alumni" in roles:
                    status = "Alumni"
                else:
                    status = "Member"

                # Determine piano-playing group based on roles
                pg = ", ".join(r for r in roles if r in {"Advanced","Intermediate","Novice","Foundational"}) or "None"
                rows.append({
                    'Discord_Username': m.name,
                    'Name': m.nick or "None",
                    'Role': status,
                    'Piano_Playing_Group': pg,
                    'Joined_Server_Time': m.joined_at.astimezone(SGT).strftime("%Y-%m-%d %H:%M:%S")
                })

        if not rows:
            logger.warning("No members and alumni found for Excel export.")
            await interaction.response.send_message("No members and alumni found.")
            return

        df = pd.DataFrame(rows)
        fname = "data/members_details.xlsx"
        df.to_excel(fname, index=False)

        # Load the workbook and convert the sheet to a table
        wb = load_workbook(fname)
        ws = wb.active

        # Define the table range and name
        tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
        tbl = Table(displayName="MemberTable", ref=tab_ref)

        # Add style to the table
        style = TableStyleInfo(name="TableStyleLight18", showRowStripes=True)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        wb.save(fname)
        logger.info("Excel table formatting complete.")

        await interaction.followup.send("Here is the Excel export:", file=discord.File(fname))
        os.remove(fname)
        logger.info("Temporary Excel file removed.")


    @exco_exclusive_group.command(name="weekly-session-nominal-rolls", description="Exports nominal rolls of all weekly sessions to Excel.")
    @has_allowed_role_and_channel(allowed_roles=['Current EXCO', 'Admin'], allowed_channels=['👑┃exco-exclusive', '⚙️┃admin-related'])
    async def weekly_session_nominal_rolls(self, interaction: discord.Interaction):

        await interaction.response.defer()

        log_slash_command(logger, interaction)

        df = pd.read_csv('data/all_bookings.csv')
        fname = 'data/all_bookings.xlsx'
        df.to_excel(fname, index=False)

        # Load the workbook and convert the sheet to a table
        wb = load_workbook(fname)
        ws = wb.active

        # Define the table range and name
        tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
        tbl = Table(displayName="NominalRollsTable", ref=tab_ref)

        # Add style to the table
        style = TableStyleInfo(name="TableStyleLight18", showRowStripes=True)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        wb.save(fname)
        logger.info("Excel table formatting complete.")

        await interaction.followup.send("Here is the Excel export:", file=discord.File(fname))
        os.remove(fname)
        logger.info("Temporary Excel file removed.")


async def setup(bot: commands.Bot):
    await bot.add_cog(EXCOExclusive(bot), guild=Object(id=GUILD_ID))