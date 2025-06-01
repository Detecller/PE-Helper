import discord
from discord.ext import commands
from discord import app_commands, Object
from discord.ui import Select, View
from utils.permissions import has_allowed_role_and_channel
from utils.variables import SGT
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
from utils.setup_logger import log_slash_command
import logging


GUILD_ID = int(os.getenv("GUILD_ID"))

# Get logger
logger = logging.getLogger("pe_helper")


class Members(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot


    @app_commands.command(name="list_current_exco", description="Lists names of those in the current EXCO.")
    @has_allowed_role_and_channel()
    async def list_current_exco(self, interaction: discord.Interaction):

        log_slash_command(logger, interaction)

        guild = interaction.guild
        exco = sorted([m.display_name for m in guild.members if any(r.name=="Current EXCO" for r in m.roles)], key=str.lower)
        
        if not exco:
            logger.warning("No current EXCO members found.")
            await interaction.response.send_message("No current EXCO members found.")
            return
        
        logger.info(f"Found {len(exco)} EXCO members.")
        text = "**Names of Current EXCO Members:**\n" + "\n".join(f"- {n}" for n in exco)
        await interaction.response.send_message(text)

    
    @app_commands.command(name="members_details", description="Exports member & alumni details to Excel.")
    @has_allowed_role_and_channel(forbidden_roles=['Member','Alumni'], forbidden_channels=['💬┃general'])
    async def members_details(self, interaction: discord.Interaction):

        log_slash_command(logger, interaction)
        
        guild = interaction.guild
        rows = []
        for m in guild.members:
            if m.bot: continue
            roles = {r.name for r in m.roles}

            # Assign role conditionally
            if roles & {"Member","Alumni","Current EXCO"}:
                if "Current EXCO" in roles:
                    status = "Current EXCO"
                elif "Alumni" in roles:
                    status = "Alumni"
                else:
                    status = "Member"

                # Determine piano-playing group based on roles
                pg = ", ".join(r for r in roles if r in {"Advanced","Intermediate","Novice","Foundational"}) or "None"
                rows.append({
                    'Discord_Username': m.name,
                    'Name': m.nick or "None",
                    'Role': status,
                    'Piano_Playing_Group': pg,
                    'Joined_Server_Time': m.joined_at.astimezone(SGT).strftime("%Y-%m-%d %H:%M:%S")
                })

        if not rows:
            logger.warning("No members and alumni found for Excel export.")
            await interaction.response.send_message("No members and alumni found.")
            return

        df = pd.DataFrame(rows)
        fname = "data/members_details.xlsx"
        df.to_excel(fname, index=False)

        # Load the workbook and convert the sheet to a table
        wb = load_workbook(fname)
        ws = wb.active

        # Define the table range and name
        tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
        tbl = Table(displayName="MemberTable", ref=tab_ref)

        # Add style to the table
        style = TableStyleInfo(name="TableStyleLight18", showRowStripes=True)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        wb.save(fname)
        logger.info("Excel table formatting complete.")

        await interaction.response.send_message("Here is the Excel export:", file=discord.File(fname))
        os.remove(fname)
        logger.info("Temporary Excel file removed.")

    
    @app_commands.command(name="list_piano_group_members", description="Select a piano group and list its members (excl. alumni).")
    @has_allowed_role_and_channel()
    async def list_piano_group_members(self, interaction: discord.Interaction):

        log_slash_command(logger, interaction)

        class Dropdown(Select):
            def __init__(self):

                # Create dropdown options for each piano group
                opts = [discord.SelectOption(label=g) for g in ["Foundational","Novice","Intermediate","Advanced"]]
                super().__init__(placeholder="Choose a group…", min_values=1, max_values=1, options=opts)


            async def callback(self, inter: discord.Interaction):
                grp = self.values[0]
                logger.info(f"Piano group selected: {grp} by {inter.user.display_name}")

                # Get list of members from the selected group (excluding alumni)
                names = [m.display_name for m in inter.guild.members if grp in {r.name for r in m.roles} and "Member" in {r.name for r in m.roles}]
                
                # If there are members in the group, sort and display their names
                if names:
                    names.sort(key=str.lower)
                    out = "\n".join(f"- {n}" for n in names)
                    logger.info(f"Listed {len(names)} members in group {grp}.")
                    await inter.response.send_message(f"**{grp} members:**\n{out}", ephemeral=True)
                else:
                    logger.warning(f"No current members found in group {grp}.")
                    await inter.response.send_message(f"No current members in {grp}.", ephemeral=True)


        # Create a view for the select dropdown
        view = View()
        view.add_item(Dropdown())
        
        logger.info("Dropdown sent for piano group selection.")
        await interaction.response.send_message("Please select a group:", view=view, ephemeral=True)


    @app_commands.command(name="weekly_session_nominal_rolls", description="Exports nominal rolls of all weekly sessions to Excel.")
    @has_allowed_role_and_channel(forbidden_roles=['Member','Alumni'], forbidden_channels=['💬┃general'])
    async def weekly_session_nominal_rolls(self, interaction: discord.Interaction):

        log_slash_command(logger, interaction)

        df = pd.read_csv('data/all_bookings.csv')
        fname = 'data/all_bookings.xlsx'
        df.to_excel(fname, index=False)

        # Load the workbook and convert the sheet to a table
        wb = load_workbook(fname)
        ws = wb.active

        # Define the table range and name
        tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
        tbl = Table(displayName="NominalRollsTable", ref=tab_ref)

        # Add style to the table
        style = TableStyleInfo(name="TableStyleLight18", showRowStripes=True)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(c.value)) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        wb.save(fname)
        logger.info("Excel table formatting complete.")

        await interaction.response.send_message("Here is the Excel export:", file=discord.File(fname))
        os.remove(fname)
        logger.info("Temporary Excel file removed.")


async def setup(bot: commands.Bot):
    await bot.add_cog(Members(bot), guild=Object(id=GUILD_ID))