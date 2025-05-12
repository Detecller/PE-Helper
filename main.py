import discord
from discord.ext import tasks
from discord import app_commands
from discord.ui import Select, View
import pandas as pd
from dotenv import load_dotenv
import os
import plotly.graph_objects as go
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import datetime
import pytz

load_dotenv()
TOKEN = os.getenv('DISCORD_TOKEN')

intents = discord.Intents.default()
intents.members = True
intents.guilds = True
intents.messages = True
intents.message_content = True

bot = discord.Client(intents=intents)
tree = app_commands.CommandTree(bot)

SGT = pytz.timezone('Asia/Singapore')

# Check Permissions
def has_allowed_role_and_channel(forbidden_roles: list[str] = None, forbidden_channels: list[str] = None):
    # Default allowed roles and channels
    default_allowed_roles = {'Admin', 'Current EXCO', 'Member', 'Alumni'}
    default_allowed_channels = {'🛠┃admin-discussions', '🧠┃exco-discussions', '💬┃general', '🤖┃bot-dev'}

    # Ensure the parameters are not None and set defaults if empty
    forbidden_roles = forbidden_roles or []
    forbidden_channels = forbidden_channels or []

    # Validate user roles and channel
    async def predicate(interaction: discord.Interaction) -> bool:

        # Calculate the allowed roles and channels by removing forbidden ones
        allowed_roles = default_allowed_roles - set(forbidden_roles)
        allowed_channels = default_allowed_channels - set(forbidden_channels)

        user_roles = {r.name for r in interaction.user.roles}
        channel_name = interaction.channel.name if interaction.channel else None

        # Check if user has at least one of the allowed roles
        if not (user_roles & allowed_roles):
            await interaction.response.send_message("❌ You don't have the permitted role.", ephemeral=True)
            return False

        # Check if the channel is allowed
        if channel_name not in allowed_channels:
            await interaction.response.send_message("❌ This command can't be used in this channel.", ephemeral=True)
            return False

        return True

    return app_commands.check(predicate)


@bot.event
async def on_ready():
    print(f"Logged in as {bot.user} (ID: {bot.user.id})")

    # Sync all slash commands
    guild = discord.utils.get(bot.guilds, name="NYP Piano Ensemble")
    if guild:
        await tree.sync(guild=discord.Object(id=guild.id))

    if not count_messages.is_running():
        count_messages.start()


@bot.event
async def on_app_command_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    if isinstance(error, app_commands.CheckFailure):
        return
    raise error


@tasks.loop(hours=1)
async def count_messages():
    guild = discord.utils.get(bot.guilds, name="NYP Piano Ensemble")
    message_counts: dict[str, int] = {}
    word_counts: dict[str, int] = {}
    target_roles = ['Member', 'Alumni']
    role_objs = [discord.utils.get(guild.roles, name=r) for r in target_roles]

    scanned = []
    for ch in guild.text_channels:
        if not any(ch.permissions_for(role).view_channel for role in role_objs if role):
            continue
        try:
            async for msg in ch.history(limit=None):
                if not isinstance(msg.author, discord.Member) or msg.author.bot:
                    continue
                name = msg.author.display_name
                message_counts.setdefault(name, 0)
                word_counts.setdefault(name, 0)
                message_counts[name] += 1
                word_counts[name] += len(msg.content.split())
        except discord.Forbidden:
            continue
        scanned.append(ch.name)

    df = pd.DataFrame([
        {"Name": n, "Message Count": message_counts[n], "Word Count": word_counts[n]}
        for n in message_counts
    ])
    df.sort_values("Message Count", ascending=False).head(10).to_csv("top_messages.csv", index=False)
    df.sort_values("Word Count", ascending=False).head(10).to_csv("top_words.csv", index=False)
    with open("channels.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(scanned))
    global last_update
    last_update = datetime.datetime.now(SGT)


@tree.command(name="piano_groups", description="Pie chart of piano-playing groups of current members.")
@has_allowed_role_and_channel()
async def piano_groups(interaction: discord.Interaction):
    guild = interaction.guild
    count_dict = {"Advanced": 0, "Intermediate": 0, "Novice": 0, "Foundational": 0}

    for m in guild.members:
        if m.bot:
            continue
        roles = {r.name for r in m.roles}
        if "Member" in roles:  # Obtain current members only
            if "Advanced" in roles:
                count_dict["Advanced"] += 1
            elif "Intermediate" in roles:
                count_dict["Intermediate"] += 1
            elif "Novice" in roles:
                count_dict["Novice"] += 1
            elif "Foundational" in roles:
                count_dict["Foundational"] += 1

    # Generate pie chart for roles
    labels = ["Foundational", "Novice", "Intermediate", "Advanced"]
    values = [count_dict[label] for label in labels]

    label_colors = {
        "Advanced": "#05668d",
        "Intermediate": "#427aa1",
        "Novice": "#679436",
        "Foundational": "#a5be00"
    }

    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        marker=dict(colors=[label_colors[label] for label in labels]),
        textinfo='value+percent',
        sort=False
    )])
    fig.update_layout(
        title="Piano-Playing Groups of Current Members",
        legend=dict(traceorder="normal")
    )

    # Save the figure to an image in memory
    buf = BytesIO()
    fig.write_image(buf, format="png")
    buf.seek(0)

    await interaction.response.send_message(file=discord.File(buf, "piano_groups.png"))


@tree.command(name="list_current_exco", description="Lists names of those in the current EXCO.")
@has_allowed_role_and_channel()
async def list_current_exco(interaction: discord.Interaction):
    guild = interaction.guild
    exco = sorted([m.display_name for m in guild.members if any(r.name=="Current EXCO" for r in m.roles)], key=str.lower)
    text = "**Names of Current EXCO Members:**\n" + "\n".join(f"- {n}" for n in exco)
    await interaction.response.send_message(text)


@tree.command(name="list_piano_group_members", description="Select a piano group and list its members (excl. alumni).")
@has_allowed_role_and_channel()
async def list_piano_group_members(interaction: discord.Interaction):

    class Dropdown(Select):
        def __init__(self):

            # Create dropdown options for each piano group
            opts = [discord.SelectOption(label=g) for g in ["Foundational","Novice","Intermediate","Advanced"]]
            super().__init__(placeholder="Choose a group…", min_values=1, max_values=1, options=opts)


        async def callback(self, inter: discord.Interaction):
            grp = self.values[0]

            # Get list of members from the selected group (excluding alumni)
            names = [m.display_name for m in inter.guild.members if grp in {r.name for r in m.roles} and "Member" in {r.name for r in m.roles}]
            
            # If there are members in the group, sort and display their names
            if names:
                names.sort(key=str.lower)
                out = "\n".join(f"- {n}" for n in names)
                await inter.response.send_message(f"**{grp} members:**\n{out}", ephemeral=True)
            else:
                await inter.response.send_message(f"No current members in {grp}.", ephemeral=True)


    # Create a view for the select dropdown
    view = View()
    view.add_item(Dropdown())
    
    await interaction.response.send_message("Please select a group:", view=view, ephemeral=True)


@tree.command(name="message_stats", description="Bar charts of total messages & word counts by user.")
@has_allowed_role_and_channel()
async def message_stats(interaction: discord.Interaction):
    with open("channels.txt", "r", encoding="utf-8") as f:
        channels = f.read().splitlines()
    
    df_msg = pd.read_csv("top_messages.csv")
    df_words = pd.read_csv("top_words.csv")

    # Create a horizontal bar chart of message counts
    fig1 = go.Figure(data=go.Bar(
        x=df_msg["Message Count"],
        y=df_msg["Name"],
        orientation='h',
        text=df_msg["Message Count"],
        marker=dict(color='#1985a1')
    ))
    fig1.update_layout(
        title='Top 10 Message Senders',
        xaxis_title='Total Messages',
        yaxis_title='Name',
        yaxis=dict(autorange='reversed', ticksuffix='  '),
        plot_bgcolor='white',
        title_x=0.5
    )
    buf1 = BytesIO()
    fig1.write_image(buf1, format="png")
    buf1.seek(0)

    # Create a horizontal bar chart of word counts
    fig2 = go.Figure(data=go.Bar(
        x=df_words["Word Count"],
        y=df_words["Name"],
        orientation='h',
        text=df_words["Word Count"],
        marker=dict(color='#284b63')
    ))
    fig2.update_layout(
        title='Top 10 Users by Word Count',
        xaxis_title='Total Words',
        yaxis_title='Name',
        yaxis=dict(autorange='reversed', ticksuffix='  '),
        plot_bgcolor='white',
        title_x=0.5
    )
    buf2 = BytesIO()
    fig2.write_image(buf2, format="png")
    buf2.seek(0)

    # Send channel list header and the images
    header = "**Scanned Channels:**\n" + "\n".join(f"- {c}" for c in channels)
    await interaction.response.send_message(header)
    await interaction.followup.send(file=discord.File(buf1, "message_count.png"))
    await interaction.followup.send(file=discord.File(buf2, "word_count.png"))
    await interaction.followup.send(f"Last updated: {last_update.strftime('%Y-%m-%d %H:%M:%S')} SGT")


@tree.command(name="members_details", description="Exports member & alumni details to Excel.")
@has_allowed_role_and_channel(forbidden_roles=['Member','Alumni'], forbidden_channels=['💬┃general'])
async def members_details(interaction: discord.Interaction):
    guild = interaction.guild
    rows = []
    for m in guild.members:
        if m.bot: continue
        roles = {r.name for r in m.roles}

        # Assign role conditionally
        if roles & {"Member","Alumni","Current EXCO"}:
            if "Current EXCO" in roles:
                status = "Current EXCO"
            elif "Alumni" in roles:
                status = "Alumni"
            else:
                status = "Member"

            # Determine piano-playing group based on roles
            pg = ", ".join(r for r in roles if r in {"Advanced","Intermediate","Novice","Foundational"}) or "None"
            rows.append({
                'Discord_Username': m.name,
                'Name': m.nick or "None",
                'Role': status,
                'Piano_Playing_Group': pg,
                'Joined_Server_Time': m.joined_at.astimezone(SGT).strftime("%Y-%m-%d %H:%M:%S")
            })

    df = pd.DataFrame(rows)
    fname = "members_details.xlsx"
    df.to_excel(fname, index=False)

    # Load the workbook and convert the sheet to a table
    wb = load_workbook(fname)
    ws = wb.active

    # Define the table range and name
    tab_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    tbl = Table(displayName="MemberTable", ref=tab_ref)

    # Add style to the table
    style = TableStyleInfo(name="TableStyleLight18", showRowStripes=True)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    wb.save(fname)

    await interaction.response.send_message("Here is the Excel export:", file=discord.File(fname))
    os.remove(fname)


bot.run(TOKEN)